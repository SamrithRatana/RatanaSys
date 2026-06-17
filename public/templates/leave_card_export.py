#!/usr/bin/env python3
"""
leave_card_export.py
Reads JSON payload from stdin, writes xlsx bytes to stdout.
Place this file at: public/scripts/leave_card_export.py

JSON input shape:
{
  "templatePath": "/abs/path/to/leave-card.xlsx",
  "headerName":   "ឈ្មោះបុគ្គលិក៖  Name",
  "headerPos":    "តួនាទី៖  Position",
  "headerDept":   "ផ្នែក/សាខា  Dept",
  "headerCredit": "ច្បាប់ឈប់សម្រាកប្រចាំឆ្នាំរយៈពេល ១៨ ថ្ងៃ",
  "sec1": [{"applied":"dd/mm/yyyy","start":"...","end":"...","dur":"...","balance":"...","note":"..."}],
  "sec2": [...],  // sick — note goes to col 7
  "sec3": [...]   // special/maternity
}
"""
import sys, json, copy
from io import BytesIO
import openpyxl
from openpyxl import load_workbook


def copy_row_style(ws, src_row: int, dst_row: int):
    """Copy formatting (not values) from src_row to dst_row."""
    src_rd = ws.row_dimensions.get(src_row)
    if src_rd and src_rd.height:
        ws.row_dimensions[dst_row].height = src_rd.height
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font      = copy.copy(src_cell.font)
            dst_cell.fill      = copy.copy(src_cell.fill)
            dst_cell.border    = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
        dst_cell.value = None  # always start blank


def insert_data_rows(ws, after_row: int, count: int, style_from_row: int):
    """
    Insert `count` blank rows immediately after `after_row`.
    openpyxl.insert_rows() correctly shifts merged cell references downward,
    avoiding the duplication bug that occurs with ExcelJS spliceRows().
    """
    if count <= 0:
        return
    ws.insert_rows(after_row + 1, count)
    for i in range(count):
        copy_row_style(ws, style_from_row, after_row + 1 + i)


def write_row(ws, row_num: int, cols: dict):
    """Write values into specific columns of a row. cols = {col_index: value}."""
    for col, val in cols.items():
        if val is not None and val != "":
            ws.cell(row=row_num, column=col).value = val


def main():
    data = json.load(sys.stdin)
    tmpl_path = data["templatePath"]

    wb = load_workbook(tmpl_path)
    ws = wb.worksheets[0]

    # ── Fill employee header ──────────────────────────────────────────────────
    ws["A7"].value  = data.get("headerName",   "")
    ws["A8"].value  = data.get("headerPos",    "")
    ws["A9"].value  = data.get("headerDept",   "")
    ws["A11"].value = data.get("headerCredit", "")

    # ── Template data row positions (verified against leave-card.xlsx) ────────
    ANNUAL_DATA  = 15
    SICK_DATA    = 22
    SPECIAL_DATA = 28

    sec1 = data.get("sec1", [])  # annual / personal / short
    sec2 = data.get("sec2", [])  # sick
    sec3 = data.get("sec3", [])  # special / maternity

    # ── Section 1: Annual / Personal / Short ─────────────────────────────────
    extra1 = max(0, len(sec1) - 1)
    if extra1 > 0:
        insert_data_rows(ws, ANNUAL_DATA, extra1, ANNUAL_DATA)
        SICK_DATA    += extra1
        SPECIAL_DATA += extra1

    for i, lv in enumerate(sec1):
        write_row(ws, ANNUAL_DATA + i, {
            1: lv.get("applied"),
            2: lv.get("start"),
            3: lv.get("end"),
            4: lv.get("dur"),
            5: lv.get("balance"),
            6: lv.get("note"),
        })

    # ── Section 2: Sick ───────────────────────────────────────────────────────
    extra2 = max(0, len(sec2) - 1)
    if extra2 > 0:
        insert_data_rows(ws, SICK_DATA, extra2, SICK_DATA)
        SPECIAL_DATA += extra2

    for i, lv in enumerate(sec2):
        write_row(ws, SICK_DATA + i, {
            1: lv.get("applied"),
            2: lv.get("start"),
            3: lv.get("end"),
            4: lv.get("dur"),
            5: lv.get("balance"),
            7: lv.get("note"),   # sick note in col G
        })

    # ── Section 3: Special / Maternity ───────────────────────────────────────
    extra3 = max(0, len(sec3) - 1)
    if extra3 > 0:
        insert_data_rows(ws, SPECIAL_DATA, extra3, SPECIAL_DATA)

    for i, lv in enumerate(sec3):
        write_row(ws, SPECIAL_DATA + i, {
            1: lv.get("applied"),
            2: lv.get("start"),
            3: lv.get("end"),
            4: lv.get("dur"),
            5: lv.get("balance"),
            6: lv.get("note"),
        })

    # ── Write output to stdout ────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    sys.stdout.buffer.write(buf.getvalue())


if __name__ == "__main__":
    main()
